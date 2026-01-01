"""
Excel to JSON Converter - Vercel Serverless Function
"""
import json
import io
import os
import re
import hashlib
from datetime import datetime
from http.server import BaseHTTPRequestHandler
import cgi

import pandas as pd
from openpyxl import load_workbook

# ---------------- Helpers ----------------

CELL_REF_RE = re.compile(r"\$?[A-Z]{1,3}\$?\d+")
RANGE_RE = re.compile(r"(\$?[A-Z]{1,3}\$?\d+):(\$?[A-Z]{1,3}\$?\d+)")


def _hash_id(*parts):
    h = hashlib.sha1("::".join(str(p) for p in parts).encode()).hexdigest()[:4]
    return h


def _iter_non_empty(ws):
    for row in ws.iter_rows():
        for c in row:
            if c.value not in (None, ""):
                yield c


def _cell_type(v):
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return "number"
    if isinstance(v, bool):
        return "bool"
    if isinstance(v, (datetime, pd.Timestamp)):
        return "datetime"
    return "str"


def _safe_val(v):
    if isinstance(v, (datetime, pd.Timestamp)):
        return v.isoformat()
    return v


def _tuple_to_coord(rc):
    from openpyxl.utils.cell import get_column_letter
    r, c = rc
    return f"{get_column_letter(c)}{r}"


def _expand_range(rg):
    from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
    m = RANGE_RE.match(rg)
    if not m:
        return []
    a, b = m.group(1), m.group(2)
    r1, c1 = coordinate_to_tuple(a)
    r2, c2 = coordinate_to_tuple(b)
    cells = []
    for r in range(min(r1, r2), max(r1, r2) + 1):
        for c in range(min(c1, c2), max(c1, c2) + 1):
            cells.append(f"{get_column_letter(c)}{r}")
    return cells


def _formula_deps(formula):
    if not formula or not isinstance(formula, str) or not formula.startswith("="):
        return []
    refs = set()
    for m in CELL_REF_RE.finditer(formula):
        refs.add(m.group(0))
    for m in RANGE_RE.finditer(formula):
        refs.add(f"{m.group(1)}:{m.group(2)}")
    return sorted(refs)


def _detect_text_sections(ws):
    rows = {}
    for c in _iter_non_empty(ws):
        if isinstance(c.value, str) and c.value.strip():
            rows.setdefault(c.row, []).append(c)

    sections = []
    if not rows:
        return sections

    current_block_rows = []
    for r in range(1, ws.max_row + 1):
        if r in rows:
            current_block_rows.append(r)
        else:
            if current_block_rows:
                rmin, rmax = min(current_block_rows), max(current_block_rows)
                cmin = min(min(c.column for c in rows[rr]) for rr in current_block_rows)
                cmax = max(max(c.column for c in rows[rr]) for rr in current_block_rows)
                text_lines = []
                for rr in range(rmin, rmax + 1):
                    line = []
                    for cc in range(cmin, cmax + 1):
                        addr = _tuple_to_coord((rr, cc))
                        val = ws[addr].value
                        if val is not None and str(val).strip():
                            line.append(str(val))
                    if line:
                        text_lines.append(" ".join(line))
                if text_lines:
                    from openpyxl.utils.cell import get_column_letter
                    rng = f"{get_column_letter(cmin)}{rmin}:{get_column_letter(cmax)}{rmax}"
                    sections.append({"range": rng, "text": "\n".join(text_lines)})
                current_block_rows = []

    if current_block_rows:
        rmin, rmax = min(current_block_rows), max(current_block_rows)
        cmin = min(min(c.column for c in rows[rr]) for rr in current_block_rows)
        cmax = max(max(c.column for c in rows[rr]) for rr in current_block_rows)
        text_lines = []
        for rr in range(rmin, rmax + 1):
            line = []
            for cc in range(cmin, cmax + 1):
                addr = _tuple_to_coord((rr, cc))
                val = ws[addr].value
                if val is not None and str(val).strip():
                    line.append(str(val))
            if line:
                text_lines.append(" ".join(line))
        if text_lines:
            from openpyxl.utils.cell import get_column_letter
            rng = f"{get_column_letter(cmin)}{rmin}:{get_column_letter(cmax)}{rmax}"
            sections.append({"range": rng, "text": "\n".join(text_lines)})

    return sections


def _sheet_chunks(sheet_name, ws, excel_tables, sections, max_cells=400):
    chunks = []

    def chunk_id(kind, rng):
        return f"{sheet_name.lower()}_{rng.lower().replace(':', '_')}_{_hash_id(kind, rng)}"

    for t in excel_tables:
        text = f"Table {t['name']} with {len(t.get('records', []))} rows. Columns: {', '.join(t.get('headers', []))}."
        chunks.append({
            "chunk_id": chunk_id("table", t["range"]),
            "kind": "table",
            "sheet": sheet_name,
            "range": t["range"],
            "text": text,
            "cells": _expand_range(t["range"])[:max_cells]
        })

    for s in sections:
        chunks.append({
            "chunk_id": chunk_id("section", s["range"]),
            "kind": "section",
            "sheet": sheet_name,
            "range": s["range"],
            "text": s["text"],
            "cells": _expand_range(s["range"])[:max_cells]
        })

    return chunks


def _open_workbooks(file_bytes):
    b1 = io.BytesIO(file_bytes)
    wb_values = load_workbook(b1, data_only=True)
    b2 = io.BytesIO(file_bytes)
    wb_formulas = load_workbook(b2, data_only=False)
    return wb_values, wb_formulas


def _extract_named_ranges(wb):
    named = []
    try:
        iterable = getattr(wb.defined_names, "definedName", None)
        if iterable is None:
            iterable = wb.defined_names
        for dn in iterable:
            try:
                if not hasattr(dn, "name"):
                    dn = wb.defined_names[dn]
            except Exception:
                pass
            if getattr(dn, "name", "").startswith("_xlnm."):
                continue
            destinations = []
            try:
                if hasattr(dn, "destinations"):
                    destinations = list(dn.destinations)
            except Exception:
                destinations = []
            if destinations:
                for sheet, ref in destinations:
                    named.append({"name": dn.name, "sheet": sheet, "ref": ref})
            else:
                raw = getattr(dn, "attr_text", None)
                named.append({"name": getattr(dn, "name", None), "sheet": None, "ref": raw})
    except Exception as e:
        named.append({"error": f"named-range-parse: {e}"})
    return named


def extract_excel_ai(
    file_bytes: bytes,
    include_formulas: bool = True,
    include_cells: bool = True,
    include_comments: bool = True,
    include_named_ranges: bool = True,
    include_excel_tables: bool = True,
    include_inferred_sections: bool = True,
    chunk_max_cells: int = 400,
    file_name: str = "workbook"
):
    wb_values, wb_formula = _open_workbooks(file_bytes)

    meta = {
        "title": file_name,
        "sheets": wb_values.sheetnames,
        "created": _safe_val(getattr(wb_values.properties, "created", None)),
        "modified": _safe_val(getattr(wb_values.properties, "modified", None))
    }

    out = {"workbook": meta, "sheets": {}}

    if include_named_ranges:
        out["named_ranges"] = _extract_named_ranges(wb_values)

    for sheet in wb_values.sheetnames:
        ws = wb_values[sheet]
        ws_f = wb_formula[sheet]

        sheet_obj = {}
        sheet_obj["dims"] = ws.calculate_dimension()
        sheet_obj["frozen_panes"] = getattr(ws, "freeze_panes", None)
        sheet_obj["merged_ranges"] = [str(rng) for rng in ws.merged_cells.ranges]

        if include_cells:
            cells = {}
            for c in _iter_non_empty(ws):
                addr = c.coordinate
                val = _safe_val(c.value)
                t = _cell_type(c.value)
                item = {"v": val, "t": t, "display": str(c.value) if c.value is not None else ""}
                if include_formulas:
                    c_f = ws_f[addr]
                    fv = c_f.value if (isinstance(c_f.value, str) and str(c_f.value).startswith("=")) else None
                    if fv:
                        item["f"] = fv
                        item["deps"] = _formula_deps(fv)
                if include_comments:
                    item["hyperlink"] = c.hyperlink.target if c.hyperlink else None
                    item["comment"] = c.comment.text if c.comment else None
                cells[addr] = item
            sheet_obj["cells"] = cells

        excel_tables = []
        if include_excel_tables and getattr(ws, "tables", None):
            try:
                from openpyxl.utils.cell import range_boundaries
                for name, tbl in ws.tables.items():
                    rng = tbl.ref
                    min_col, min_row, max_col, max_row = range_boundaries(rng)
                    headers = []
                    for c in ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row, max_row=min_row):
                        headers.append(str(c[0].value) if c and c[0].value is not None else "")
                    excel_tables.append({
                        "name": getattr(tbl, "displayName", None) or name,
                        "range": rng,
                        "headers": headers,
                        "records": []
                    })
            except Exception:
                pass
        sheet_obj["tables"] = excel_tables

        sections = _detect_text_sections(ws) if include_inferred_sections else []
        sheet_obj["sections"] = sections

        if include_formulas and include_cells:
            lineage_nodes = []
            for addr, item in sheet_obj.get("cells", {}).items():
                if "f" in item and item["f"]:
                    lineage_nodes.append({"cell": addr, "formula": item["f"], "deps": item.get("deps", [])})
            sheet_obj["lineage"] = {"nodes": lineage_nodes}

        chunks = _sheet_chunks(sheet, ws, excel_tables, sections, max_cells=chunk_max_cells)
        sheet_obj["chunks"] = chunks

        out["sheets"][sheet] = sheet_obj

    return out


class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_POST(self):
        try:
            content_type = self.headers.get('Content-Type', '')
            
            if 'multipart/form-data' not in content_type:
                self.send_error(400, 'Content-Type must be multipart/form-data')
                return

            # Parse multipart form data
            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    'REQUEST_METHOD': 'POST',
                    'CONTENT_TYPE': content_type,
                }
            )

            # Get file
            file_item = form['file']
            if not file_item.file:
                self.send_error(400, 'No file uploaded')
                return

            file_bytes = file_item.file.read()
            file_name = file_item.filename or 'workbook.xlsx'

            # Get options
            def get_bool(key, default=True):
                val = form.getvalue(key, str(default))
                return val.lower() == 'true'

            include_formulas = get_bool('include_formulas', True)
            include_cells = get_bool('include_cells', True)
            include_comments = get_bool('include_comments', True)
            include_named_ranges = get_bool('include_named_ranges', True)
            include_excel_tables = get_bool('include_excel_tables', True)
            include_inferred_sections = get_bool('include_inferred_sections', True)
            chunk_max_cells = int(form.getvalue('chunk_max_cells', '400'))

            # Process
            result = extract_excel_ai(
                file_bytes=file_bytes,
                include_formulas=include_formulas,
                include_cells=include_cells,
                include_comments=include_comments,
                include_named_ranges=include_named_ranges,
                include_excel_tables=include_excel_tables,
                include_inferred_sections=include_inferred_sections,
                chunk_max_cells=chunk_max_cells,
                file_name=file_name
            )

            # Send response
            response_body = json.dumps(result, default=str, ensure_ascii=False)
            
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(response_body.encode('utf-8'))

        except Exception as e:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            error_response = json.dumps({'error': str(e)})
            self.wfile.write(error_response.encode('utf-8'))

